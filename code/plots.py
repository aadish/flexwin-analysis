import matplotlib.pyplot as pyplot
from openpyxl import load_workbook

wb = load_workbook('Analysis.xlsx')
ws = wb.get_sheet_by_name('Sheet1')
content = []

for row in ws.iter_rows(row_offset=1):
    if row[0].value:
        content.append((row[0].value, int(row[1].value), int(row[2].value), int(row[3].value), int(row[4].value)))

C0 = [(content[0][4], content[0][2])]
C1 = [(content[0][4], content[0][2])]
C2 = [(content[0][4], content[0][2])]

for i in content[1:]:
    if i[0].split('=')[0] == 'C0':
        C0.append( (i[4], i[2]) )
    if i[0].split('=')[0] == 'C1':
        C1.append( (i[4], i[2]) )
    if i[0].split('=')[0] == 'C2':
        C2.append( (i[4], i[2]) )

Sorted_C0 = sorted(C0)
Sorted_C1 = sorted(C1)
Sorted_C2 = sorted(C2)

false_negative_C0 = [i[0] for i in Sorted_C0]
false_positive_C0 = [i[1] for i in Sorted_C0]

false_negative_C1 = [i[0] for i in Sorted_C1]
false_positive_C1 = [i[1] for i in Sorted_C1]

false_negative_C2 = [i[0] for i in Sorted_C2]
false_positive_C2 = [i[1] for i in Sorted_C2]


line0, = pyplot.plot(false_negative_C0, false_positive_C0, label='C0')
line1, = pyplot.plot(false_negative_C1, false_positive_C1, label='C1')
line2, = pyplot.plot(false_negative_C2, false_positive_C2, label='C2')
pyplot.legend(handles = [line0, line1, line2])
pyplot.xlabel('False Negative')
pyplot.ylabel('False Positive')
pyplot.title('False Negative vs False Positive')
pyplot.show()
